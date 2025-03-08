import os
import json
import requests
import sys
import traceback
import csv
import io
import uuid
import time
from flask import Flask, request, jsonify, send_file, Response
from langchain_community.embeddings import OllamaEmbeddings
from langchain_community.vectorstores import OpenSearchVectorSearch
from langchain.prompts import PromptTemplate
from langchain.chains import RetrievalQA
from langchain_community.document_loaders import (
    UnstructuredPDFLoader, UnstructuredImageLoader, UnstructuredWordDocumentLoader,
    UnstructuredExcelLoader, CSVLoader
)
from langchain.text_splitter import RecursiveCharacterTextSplitter
# Заменяем устаревший импорт на новый
from langchain_ollama import OllamaLLM
from redis import Redis
from tenacity import retry, stop_after_attempt, wait_exponential
import asyncpg
import logging
from openpyxl import load_workbook
from datetime import datetime
import asyncio
from collections import deque

# Налаштування логування
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')  # Исправлен формат логирования
logger = logging.getLogger(__name__)

app = Flask(__name__)

REDIS_HOST = os.getenv("REDIS_HOST", "localhost")
REDIS_PASSWORD = os.getenv("REDIS_PASSWORD")
POSTGRES_HOST = os.getenv("POSTGRES_HOST", "localhost")
POSTGRES_PASSWORD = os.getenv("POSTGRES_PASSWORD", "strong_password")
OLLAMA_HOST = os.getenv("OLLAMA_HOST", "http://host.docker.internal:11434")
OPENSEARCH_HOSTS = os.getenv("OPENSEARCH_HOSTS", "http://localhost:9200").split(",")
MAX_CONTEXT_LENGTH = int(os.getenv("MAX_CONTEXT_LENGTH", 2048))
OPENSEARCH_BULK_SIZE = int(os.getenv("OPENSEARCH_BULK_SIZE", 500))
MAX_FILE_SIZE = int(os.getenv("MAX_FILE_SIZE", 10 * 1024 * 1024))  # 10 MB за замовчуванням
API_PORT = int(os.getenv("API_PORT", 5001))  # Порт для API

logger.info(f"Connecting to Redis: {REDIS_HOST}")
redis_client = Redis(
    host=REDIS_HOST,
    port=6379,
    password=REDIS_PASSWORD if REDIS_PASSWORD else None,
    decode_responses=True
)

logger.info(f"Initializing embeddings with Ollama: {OLLAMA_HOST}")
embeddings = OllamaEmbeddings(model="nomic-embed-text", base_url=OLLAMA_HOST)

logger.info(f"Connecting to OpenSearch: {OPENSEARCH_HOSTS[0]}")
vectorstore = OpenSearchVectorSearch(
    index_name="customs_data",
    opensearch_url=OPENSEARCH_HOSTS[0],
    embedding_function=embeddings,
    http_auth=("admin", os.getenv("OPENSEARCH_PASSWORD", "strong_password")),
    use_ssl=False,
    bulk_size=OPENSEARCH_BULK_SIZE
)

RAG_PROMPT_TEMPLATE = """
You are a customs schemes and financial fraud analyst. Analyze the following data and answer the question.
Data:
{context}
Question: {question}
Answer:
"""
prompt = PromptTemplate(template=RAG_PROMPT_TEMPLATE, input_variables=["context", "question"])

@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=1, min=4, max=10))
async def get_db_pool():
    logger.info(f"Connecting to PostgreSQL: {POSTGRES_HOST}")
    return await asyncpg.create_pool(
        database="predator_db",
        user="predator",
        password=POSTGRES_PASSWORD,
        host=POSTGRES_HOST,
        command_timeout=10
    )

# Виправлена функція для індексації з буфером
async def index_csv_data(file_path, expected_headers, buffer):
    pool = await get_db_pool()
    async with pool.acquire() as conn:
        await conn.execute("TRUNCATE TABLE customs_data")
        logger.info("PostgreSQL table truncated")

        with open(file_path, 'r', encoding='utf-8') as csvfile:
            content = csvfile.read()
            csv_reader = csv.DictReader(io.StringIO(content))
            headers = csv_reader.fieldnames
            logger.info(f"CSV headers: {headers}")
            if not headers or len(headers) != len(expected_headers):
                logger.warning(f"Header mismatch, using expected headers: {headers}")
                headers = expected_headers

            rows = list(csv_reader)
            total_rows = len(rows)
            logger.info(f"Total rows in file: {total_rows}")
            buffer.append(json.dumps({"total_rows": total_rows}) + "\n")

            if not rows:
                logger.warning("No data rows found in CSV file")
                buffer.append(json.dumps({"status": "warning", "message": "No data rows found in CSV file"}) + "\n")
                await pool.close()
                return

            # Додаємо логування першого рядка для перевірки
            if rows:
                logger.info(f"First row sample: {rows[0]}")

            documents = []
            batch_size = 500
            row_count = 0

            def parse_datetime(value, default=None):
                if not value or value.strip() == "":
                    return default
                try:
                    return datetime.strptime(value, "%Y-%m-%d %H:%M:%S.%f")
                except ValueError:
                    try:
                        return datetime.strptime(value, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            return datetime.strptime(f"1970-01-01 {value}", "%Y-%м-%d %H:%М:%S")
                        except ValueError:
                            try:
                                return datetime.strptime(f"1970-01-01 {value}", "%Y-%м-%д %H:%М:%С.%f")
                            except ValueError:
                                try:
                                    value = value.replace("М", "M")
                                    return datetime.strptime(value, "%Y-%м-%д %H:%М:%С.%f")
                                except ValueError:
                                    logger.error(f"Invalid date format for value '{value}'")
                                    return default

            def parse_float(value, default=None):
                if not value or not value.strip():
                    return default
                try:
                    numeric_part = value.split()[0].replace(",", ".")
                    return float(numeric_part)
                except (ValueError, IndexError):
                    logger.warning(f"Could not parse float from '{value}', using default: {default}")
                    return default

            def parse_int(value, default=None):
                if not value or not value.strip():
                    return default
                try:
                    return int(float(value.replace(',', '.')))
                except (ValueError, TypeError):
                    logger.warning(f"Could not parse int from '{value}', using default: {default}")
                    return default

            for i, row in enumerate(rows):
                logger.info(f"Processing row {i + 1} of {total_rows}: {row}")
                try:
                    row_data = {
                        "Час оформлення": parse_datetime(row.get("Час оформлення", "")),
                        "Назва ПМО": row.get("Назва ПМО", ""),
                        "Тип": row.get("Тип", ""),
                        "Номер МД": row.get("Номер МД", ""),
                        "Дата": parse_datetime(row.get("Дата", "")),
                        "Відправник": row.get("Відправник", ""),
                        "ЕДРПОУ": parse_int(row.get("ЕДРПОУ", 0)),
                        "Одержувач": row.get("Одержувач", ""),
                        "№": parse_int(row.get("№", 0)),
                        "Код товару": row.get("Код товару", ""),
                        "Опис товару": row.get("Опис товару", ""),
                        "Кр.торг.": row.get("Кр.торг.", ""),
                        "Кр.відпр.": row.get("Кр.відпр.", ""),
                        "Кр.пох.": row.get("Кр.пох.", ""),
                        "Умови пост.": row.get("Умови пост.", ""),
                        "Місце пост": row.get("Місце пост", ""),
                        "К-ть": parse_float(row.get("К-ть", 0)) if row.get("К-ть") and row.get("К-ть").strip() else None,
                        "Один.вим.": parse_int(row.get("Один.вим.", 0)),
                        "Брутто, кг.": parse_float(row.get("Брутто, кг.", 0)) if row.get("Брутто, кг.") and row.get("Брутто, кг.").strip() else None,
                        "Нетто, кг.": parse_float(row.get("Нетто, кг.", 0)) if row.get("Нетто, кг.") and row.get("Нетто, кг.").strip() else None,
                        "Вага по МД": parse_float(row.get("Вага по МД", 0)) if row.get("Вага по МД") and row.get("Вага по МД").strip() else None,
                        "ФВ вал.контр": parse_float(row.get("ФВ вал.контр", 0)) if row.get("ФВ вал.контр") and row.get("ФВ вал.контр").strip() else None,
                        "Особ.перем.": row.get("Особ.перем.", ""),
                        "43": parse_int(row.get("43", 0)),
                        "43_01": parse_int(row.get("43_01", 0)),
                        "РФВ Дол/кг.": parse_float(row.get("РФВ Дол/кг.", 0)) if row.get("РФВ Дол/кг.") and row.get("РФВ Дол/кг.").strip() else None,
                        "Вага.один.": parse_float(row.get("Вага.один.", 0)) if row.get("Вага.один.") and row.get("Вага.один.").strip() else None,
                        "Вага різн.": parse_float(row.get("Вага різн.", 0)) if row.get("Вага різн.") and row.get("Вага різн.").strip() else None,
                        "Контракт": row.get("Контракт", ""),
                        "3001": parse_int(row.get("3001", 0)),
                        "3002": parse_int(row.get("3002", 0)),
                        "9610": parse_int(row.get("9610", 0)),
                        "Торг.марк.": row.get("Торг.марк.", ""),
                        "РМВ Нетто Дол/кг.": parse_float(row.get("РМВ Нетто Дол/кг.", 0)) if row.get("РМВ Нетто Дол/кг.") and row.get("РМВ Нетто Дол/кг.").strip() else None,
                        "РМВ Дол/дод.од.": parse_float(row.get("РМВ Дол/дод.од.", 0)) if row.get("РМВ Дол/дод.од.") and row.get("РМВ Дол/дод.од.").strip() else None,
                        "РМВ Брутто Дол/кг": parse_float(row.get("РМВ Брутто Дол/кг", 0)) if row.get("РМВ Брутто Дол/кг") and row.get("РМВ Брутто Дол/кг").strip() else None,
                        "Призн.Зед": parse_int(row.get("Призн.Зед", 0)),
                        "Мін.База Дол/кг.": parse_float(row.get("Мін.База Дол/кг.", 0)) if row.get("Мін.База Дол/кг.") and row.get("Мін.База Дол/кг.").strip() else None,
                        "Різн.мін.база": parse_float(row.get("Різн.мін.база", 0)) if row.get("Різн.мін.база") and row.get("Різн.мін.база").strip() else None,
                        "КЗ Нетто Дол/кг.": parse_float(row.get("КЗ Нетто Дол/кг.", 0)) if row.get("КЗ Нетто Дол/кг.") and row.get("КЗ Нетто Дол/кг.").strip() else None,
                        "КЗ Дол/шт.": parse_float(row.get("КЗ Дол/шт.", 0)) if row.get("КЗ Дол/шт.") and row.get("КЗ Дол/шт.").strip() else None,
                        "Різн.КЗ Дол/кг": parse_float(row.get("Різн.КЗ Дол/кг", 0)) if row.get("Різн.КЗ Дол/кг") and row.get("Різн.КЗ Дол/кг").strip() else None,
                        "Різ.КЗ Дол/шт": parse_float(row.get("Різ.КЗ Дол/шт", 0)) if row.get("Різ.КЗ Дол/шт") and row.get("Різ.КЗ Дол/шт").strip() else None,
                        "КЗ Брутто Дол/кг.": parse_float(row.get("КЗ Брутто Дол/кг.", 0)) if row.get("КЗ Брутто Дол/кг.") and row.get("КЗ Брутто Дол/кг.").strip() else None,
                        "Різ.КЗ Брутто": parse_float(row.get("Різ.КЗ Брутто", 0)) if row.get("Різ.КЗ Брутто") and row.get("Різ.КЗ Брутто").strip() else None,
                        "пільгова": parse_float(row.get("пільгова", None)) if row.get("пільгова") else None,
                        "повна": row.get("повна", "") if row.get("повна") and row.get("повна").strip() else None
                    }
                    await conn.execute("""
                        INSERT INTO customs_data (
                            "Час оформлення", "Назва ПМО", "Тип", "Номер МД", "Дата", "Відправник", "ЕДРПОУ", "Одержувач", "№",
                            "Код товару", "Опис товару", "Кр.торг.", "Кр.відпр.", "Кр.пох.", "Умови пост.", "Місце пост", "К-ть",
                            "Один.вим.", "Брутто, кг.", "Нетто, кг.", "Вага по МД", "ФВ вал.контр", "Особ.перем.", "43", "43_01",
                            "РФВ Дол/кг.", "Вага.один.", "Вага різн.", "Контракт", "3001", "3002", "9610", "Торг.марк.",
                            "РМВ Нетто Дол/кг.", "РМВ Дол/дод.од.", "РМВ Брутто Дол/кг", "Призн.Зед", "Мін.База Дол/кг.",
                            "Різн.мін.база", "КЗ Нетто Дол/кг.", "КЗ Дол/шт.", "Різн.КЗ Дол/кг", "Різ.КЗ Дол/шт",
                            "КЗ Брутто Дол/кг.", "Різ.КЗ Брутто", "пільгова", "повна"
                        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, $20,
                                  $21, $22, $23, $24, $25, $26, $27, $28, $29, $30, $31, $32, $33, $34, $35, $36, $37, $38,
                                  $39, $40, $41, $42, $43, $44, $45, $46, $47)
                    """, *row_data.values())

                    document = {
                        "content": f"{row.get('Опис товару', '')} {row.get('Відправник', '')} {row.get('Одержувач', '')} {row.get('Код товару', '')}",
                        "metadata": {"номер_мд": row.get("Номер МД")}
                    }
                    documents.append(document)

                    if len(documents) >= batch_size:
                        vectorstore.add_texts([doc["content"] for doc in documents], metadatas=[doc["metadata"] for doc in documents])
                        documents = []
                        row_count += batch_size
                        logger.info(f"Indexed {row_count} rows into OpenSearch")
                        buffer.append(json.dumps({"progress": row_count, "total": total_rows}) + "\n")
                except Exception as e:
                    logger.error(f"Error processing row {i}: {e}, row data: {row}")
                    continue

            if documents:
                vectorstore.add_texts([doc["content"] for doc in documents], metadatas=[doc["metadata"] for doc in documents])
                row_count += len(documents)
                logger.info(f"Indexed {row_count} rows into OpenSearch")
                buffer.append(json.dumps({"progress": row_count, "total": total_rows}) + "\n")

            logger.info(f"Total rows indexed: {row_count}")
            buffer.append(json.dumps({"status": "ok", "message": f"Indexed {row_count} rows into OpenSearch and PostgreSQL"}) + "\n")

    await pool.close()

# Синхронний генератор для потокового передавання
def generate_indexing(file_path, expected_headers):
    buffer = deque()
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        task = loop.create_task(index_csv_data(file_path, expected_headers, buffer))
        while not task.done() or buffer:
            if buffer:
                yield buffer.popleft()
            else:
                loop.run_until_complete(asyncio.sleep(0.1))  # Дати час асинхронній задачі виконатися
        loop.run_until_complete(task)  # Завершити задачу, якщо буфер порожній
    except Exception as e:
        logger.error(f"Error in generate_indexing: {e}\n{traceback.format_exc()}")
        yield json.dumps({"error": str(e)}) + "\n"
    finally:
        loop.close()

@app.route('/index_csv', methods=['POST'])
def index_csv():
    try:
        data = request.json
    except Exception as e:
        return jsonify({"error": f"Invalid JSON: {str(e)}"}), 400

    logger.info(f"Received request: {data}")
    if not data or "file_name" not in data:
        logger.error("File name is required")
        return Response(json.dumps({"error": "File name is required"}) + "\n", mimetype='application/x-ndjson')

    file_name = data["file_name"]
    input_path = f"/data/{file_name}"
    logger.info(f"Indexing CSV file: {input_path}")
    if not os.path.exists(input_path):
        logger.error(f"File not found: {input_path}")
        return Response(json.dumps({"error": f"File {input_path} not found"}) + "\n", mimetype='application/x-ndjson')
    if not input_path.endswith(".csv"):
        logger.error(f"Unsupported file format: {input_path}")
        return Response(json.dumps({"error": "Only .csv files are supported"}) + "\n", mimetype='application/x-ndjson')

    expected_headers = [
        "Час оформлення", "Назва ПМО", "Тип", "Номер МД", "Дата", "Відправник", "ЕДРПОУ", "Одержувач", "№",
        "Код товару", "Опис товару", "Кр.торг.", "Кр.відпр.", "Кр.пох.", "Умови пост.", "Місце пост", "К-ть",
        "Один.вим.", "Брутто, кг.", "Нетто, кг.", "Вага по МД", "ФВ вал.контр", "Особ.перем.", "43", "43_01",
        "РФВ Дол/кг.", "Вага.один.", "Вага різн.", "Контракт", "3001", "3002", "9610", "Торг.марк.",
        "РМВ Нетто Дол/кг.", "РМВ Дол/дод.од.", "РМВ Брутто Дол/кг", "Призн.Зед", "Мін.База Дол/кг.",
        "Різн.мін.база", "КЗ Нетто Дол/кг.", "КЗ Дол/шт.", "Різн.КЗ Дол/кг", "Різ.КЗ Дол/шт",
        "КЗ Брутто Дол/кг.", "Різ.КЗ Брутто", "пільгова", "повна"
    ]

    return Response(generate_indexing(input_path, expected_headers), mimetype='application/x-ndjson')

@app.route('/api/version', methods=['GET'])
def get_version():
    try:
        ollama_version_url = f"{OLLAMA_HOST}/api/version"
        response = requests.get(ollama_version_url, timeout=5)
        response.raise_for_status()
        ollama_version = response.json().get("version", "unknown")
        return Response(json.dumps({"version": ollama_version}), mimetype='application/json'), 200
    except Exception as e:
        logger.error(f"Error fetching version from Ollama: {e}")
        return Response(json.dumps({"version": "0.5.13"}), mimetype='application/json'), 200

@app.route('/api/models', methods=['GET'])
def get_models():
    try:
        logger.debug("Fetching models from Ollama")
        ollama_tags_url = f"{OLLAMA_HOST}/api/tags"
        response = requests.get(ollama_tags_url, timeout=5)
        response.raise_for_status()
        models = response.json().get("models", [])
        result = [{"id": m["name"], "name": m["name"], "object": "model"} for m in models]  # Добавлен ключ "object"
        logger.debug(f"Returning models: {json.dumps(result)}")
        return Response(json.dumps({"object": "list", "data": result}), mimetype='application/json'), 200
    except Exception as e:
        logger.error(f"Error fetching models from Ollama: {e}")
        # Возвращаем совместимый с OpenAI формат
        default_models = [{"id": "mistral:latest", "name": "Mistral Model", "object": "model"}]
        return Response(json.dumps({"object": "list", "data": default_models}), mimetype='application/json'), 200

@app.route('/api/tags', methods=['GET'])
def get_tags():
    try:
        ollama_tags_url = f"{OLLAMA_HOST}/api/tags"
        response = requests.get(ollama_tags_url, timeout=5)
        response.raise_for_status()
        models = response.json().get("models", [])
        return Response(json.dumps({
            "models": [{"name": m["name"], "model": m["name"], "modified_at": m.get("modified_at", "unknown"), "size": m.get("size", 0), "digest": m.get("digest", "unknown")} for m in models]
        }), mimetype='application/json'), 200
    except Exception as e:
        logger.error(f"Error fetching tags from Ollama: {e}")
        return Response(json.dumps({
            "models": [
                {"name": "mistral:latest", "model": "mistral:latest", "modified_at": "2025-03-05T20:00:00Z", "size": 0, "digest": "sha256:placeholder"},
                {"name": "custom:latest", "model": "custom:latest", "modified_at": "2025-03-05T20:00:00Z", "size": 0, "digest": "sha256:placeholder"}
            ]
        }), mimetype='application/json'), 200

def is_json_string(text):
    """Проверяет, является ли текст JSON-строкой."""
    text = text.strip()
    if text.startswith('{') and text.endswith('}'):
        try:
            json.loads(text)
            return True
        except json.JSONDecodeError:
            pass
    return False

def format_json_response(json_text):
    """Форматирует JSON-ответ в читаемый текст."""
    try:
        parsed_json = json.loads(json_text)
        if "tags" in parsed_json:
            tags = ", ".join(parsed_json.get("tags", []))
            subtopics = ", ".join(parsed_json.get("subtopics", []))
            return f"Категории: {tags}\nПодкатегории: {subtopics}"
        else:
            # Используем pretty print для других JSON структур
            return json.dumps(parsed_json, indent=2, ensure_ascii=False)
    except json.JSONDecodeError:
        return json_text  # Возвращаем исходный текст, если не удалось распарсить

@app.after_request
def after_request(response):
    """Добавляет CORS заголовки ко всем ответам."""
    response.headers.add('Access-Control-Allow-Origin', '*')
    response.headers.add('Access-Control-Allow-Headers', 'Content-Type,Authorization')
    response.headers.add('Access-Control-Allow-Methods', 'GET,PUT,POST,DELETE,OPTIONS')
    return response

@app.route('/api/chat', methods=['POST'])
def chat():
    try:
        data = request.json
        logger.debug(f"Received chat request: {json.dumps(data)}")
        
        if not data or not isinstance(data.get("messages"), list) or not data["messages"]:
            return jsonify({"error": "Messages must be a non-empty list"}), 400

        query = data["messages"][-1].get("content", "")
        model = data.get("model", "mistral:latest")
        stream = data.get('stream', False)
        limit = int(data.get("limit", 5))
        offset = int(data.get("offset", 0))

        if not query:
            return jsonify({"error": "Query content is required"}), 400

        cache_key = f"chat_cache:{query}:{limit}:{offset}:{model}"
        if cached := redis_client.get(cache_key):
            if stream:
                logger.debug(f"Returning cached stream response for query: {query}")
                return Response(cached, mimetype='application/x-ndjson')
            else:
                try:
                    response_data = json.loads(cached)
                    logger.debug(f"Returning cached response: {json.dumps(response_data)}")
                    return jsonify(response_data)
                except Exception as e:
                    logger.error(f"Error processing cached response: {e}")
        
        # Получение документов из OpenSearch
        docs = vectorstore.similarity_search(query, k=limit, offset=offset)
        
        # Получение данных из PostgreSQL
        async def fetch_results():
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                return await conn.fetch("""
                    SELECT "Номер МД", "Опис товару", "Відправник", "Одержувач", "Код товару"
                    FROM customs_data
                    WHERE "Опис товару" ILIKE $1 OR "Відправник" ILIKE $1 OR "Одержувач" ILIKE $1 OR "Код товару" ILIKE $1
                    LIMIT $2 OFFSET $3
                """, f"%{query}%", limit, offset)

        sql_results = asyncio.run(fetch_results())
        
        # Объединение и сокращение контекста
        context = "\n".join(doc.page_content for doc in docs) + "\nSQL Results:\n" + "\n".join(str(row) for row in sql_results)
        truncated_context = context[:MAX_CONTEXT_LENGTH] if len(context) > MAX_CONTEXT_LENGTH else context
        
        # Используем новый API LangChain для работы с Ollama
        llm = OllamaLLM(model=model, base_url=OLLAMA_HOST)
        qa_chain = RetrievalQA.from_chain_type(
            llm=llm,
            retriever=vectorstore.as_retriever(),
            return_source_documents=True,
            chain_type_kwargs={"prompt": prompt}
        )
        
        # Используем invoke вместо прямого вызова
        response = qa_chain.invoke({"query": query, "context": truncated_context})
        answer = response["result"]
        logger.debug(f"Generated answer: {answer}")
        
        # Проверяем, является ли ответ JSON-строкой и форматируем его
        if is_json_string(answer):
            formatted_answer = format_json_response(answer)
            logger.debug(f"Formatted JSON response: {formatted_answer}")
        else:
            formatted_answer = answer
        
        if stream:
            def generate():
                chat_id = f"chat-{uuid.uuid4()}"
                created_time = int(time.time())
                
                # Отправляем начальный чанк с ролью
                initial_chunk = {
                    "id": chat_id,
                    "object": "chat.completion.chunk",
                    "created": created_time,
                    "model": model,
                    "choices": [{
                        "index": 0,
                        "delta": {"role": "assistant"},
                        "finish_reason": None
                    }]
                }
                yield json.dumps(initial_chunk) + "\n"
                
                # Разбиваем ответ на части
                # Используем более крупные части для уменьшения задержек и проблем с разбивкой UTF-8 символов
                chunk_size = 15  # слов в одном чанке
                words = formatted_answer.split()
                chunks = [' '.join(words[i:i+chunk_size]) for i in range(0, len(words), chunk_size)]
                
                for chunk_text in chunks:
                    chunk = {
                        "id": chat_id,
                        "object": "chat.completion.chunk",
                        "created": created_time,
                        "model": model,
                        "choices": [{
                            "index": 0,
                            "delta": {"content": chunk_text + " "},
                            "finish_reason": None
                        }]
                    }
                    yield json.dumps(chunk) + "\n"
                
                # Финальный чанк для обозначения конца стрима
                final_chunk = {
                    "id": chat_id,
                    "object": "chat.completion.chunk",
                    "created": created_time,
                    "model": model,
                    "choices": [{
                        "index": 0,
                        "delta": {},
                        "finish_reason": "stop"
                    }]
                }
                yield json.dumps(final_chunk) + "\n"
            
            # Создаем и собираем ответы для кэширования
            response_generator = generate()
            response_chunks = list(response_generator)
            full_response = "".join(response_chunks)
            
            # Сохраняем в кэш
            redis_client.setex(cache_key, 3600, full_response)
            logger.debug(f"Cached stream response for key: {cache_key}")
            
            # Возвращаем полностью сформированный поток чанков
            return Response((chunk for chunk in response_chunks), mimetype='application/x-ndjson')
        else:
            # Обычный (не потоковый) ответ
            response_data = {
                "id": f"chat-{uuid.uuid4()}",
                "object": "chat.completion",
                "created": int(time.time()),
                "model": model,
                "choices": [{
                    "index": 0,
                    "message": {
                        "role": "assistant", 
                        "content": formatted_answer
                    },
                    "finish_reason": "stop"
                }]
            }
            
            # Сохранить в кэше
            redis_client.setex(cache_key, 3600, json.dumps(response_data))
            logger.debug(f"Cached non-stream response for key: {cache_key}")
            logger.debug(f"Returning response: {json.dumps(response_data)}")
            return jsonify(response_data)
    
    except Exception as e:
        logger.error(f"Error in chat endpoint: {e}\n{traceback.format_exc()}")
        return jsonify({"error": str(e), "detail": "Internal server error in chat endpoint"}), 500

@app.route('/process_query', methods=['POST'])
def process_query():
    try:
        data = request.json
        if not data or "query" not in data:
            return Response(json.dumps({"error": "Query is required"}), mimetype='application/json'), 400

        query = data["query"]
        model = data.get("model", "mistral:latest")
        limit = int(data.get("limit", 5))
        offset = int(data.get("offset", 0))

        cache_key = f"query_cache:{query}:{limit}:{offset}:{model}"
        if cached := redis_client.get(cache_key):
            return Response(json.dumps({"source": "cache", "response": json.loads(cached)}), mimetype='application/json'), 200

        docs = vectorstore.similarity_search(query, k=limit, offset=offset)

        async def fetch_results():
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                return await conn.fetch("""
                    SELECT "Номер МД", "Опис товару", "Відправник", "Одержувач", "Код товару"
                    FROM customs_data
                    WHERE "Опис товару" ILIKE $1 OR "Відправник" ILIKE $1 OR "Одержувач" ILIKE $1 OR "Код товару" ILIKE $1
                    LIMIT $2 OFFSET $3
                """, f"%{query}%", limit, offset)

        sql_results = asyncio.run(fetch_results())

        if not sql_results and not docs:
            return Response(json.dumps({"warning": "No results found in SQL or OpenSearch"}), mimetype='application/json'), 200

        context = "\n".join(doc.page_content for doc in docs) + "\nSQL Results:\n" + "\n".join(str(row) for row in sql_results)
        truncated_context = context[:MAX_CONTEXT_LENGTH] if len(context) > MAX_CONTEXT_LENGTH else context
        llm = Ollama(model=model, base_url=OLLAMA_HOST)
        qa_chain = RetrievalQA.from_chain_type(
            llm=llm,
            retriever=vectorstore.as_retriever(),
            return_source_documents=True,
            chain_type_kwargs={"prompt": prompt}
        )
        response = qa_chain({"query": query, "context": truncated_context})

        final_response = {
            "answer": response["result"],
            "sources": [doc.metadata for doc in docs] + [{"sql": dict(row)} for row in sql_results]
        }
        redis_client.setex(cache_key, 3600, json.dumps(final_response))

        async def log():
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                await conn.execute("INSERT INTO query_log (query, response) VALUES ($1, $2)", query, final_response["answer"])
        asyncio.run(log())

        return Response(json.dumps({"source": "rag", "response": final_response}), mimetype='application/json'), 200
    except Exception as e:
        logger.error(f"Error in process_query: {e}\n{traceback.format_exc()}")
        return Response(json.dumps({"error": str(e)}), mimetype='application/json'), 500

@app.route('/upload_document', methods=['POST'])
def upload_document():
    try:
        if 'file' not in request.files:
            return Response(json.dumps({"error": "File not provided"}), mimetype='application/json'), 400

        file = request.files['file']
        if file.content_length > MAX_FILE_SIZE:
            return Response(json.dumps({"error": f"File size exceeds {MAX_FILE_SIZE / (1024 * 1024)} MB"}), mimetype='application/json'), 400

        filename = file.filename.lower()
        temp_path = f"/tmp/{filename}"
        file.save(temp_path)

        if filename.endswith(".pdf"):
            loader = UnstructuredPDFLoader(temp_path)
        elif filename.endswith((".jpeg", ".jpg", ".png")):
            loader = UnstructuredImageLoader(temp_path)
        elif filename.endswith(".docx"):
            loader = UnstructuredWordDocumentLoader(temp_path)
        elif filename.endswith((".xls", ".xlsx")):
            loader = UnstructuredExcelLoader(temp_path)
        elif filename.endswith(".csv"):
            loader = CSVLoader(temp_path)
        else:
            return Response(json.dumps({"error": "Unsupported file format"}), mimetype='application/json'), 400

        try:
            documents = loader.load()
            text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)
            split_docs = text_splitter.split_documents(documents)
            vectorstore.add_texts([doc.page_content for doc in split_docs], metadatas=[{"filename": filename} for _ in split_docs])
        finally:
            os.remove(temp_path)

        return Response(json.dumps({"status": "ok", "message": f"File {filename} successfully processed"}), mimetype='application/json'), 200
    except Exception as e:
        logger.error(f"Error in upload_document: {e}\n{traceback.format_exc()}")
        return Response(json.dumps({"error": str(e)}), mimetype='application/json'), 500

@app.route('/convert_excel_to_csv_from_data', methods=['POST'])
def convert_excel_to_csv_from_data():
    try:
        data = request.json
        logger.info(f"Received request: {data}")
        if not data or "file_name" not in data:
            logger.error("File name is required")
            return Response(json.dumps({"error": "File name is required"}), mimetype='application/json'), 400

        file_name = data["file_name"]
        input_path = f"/data/{file_name}"
        logger.info(f"Checking file: {input_path}")
        if not os.path.exists(input_path):
            logger.error(f"File not found: {input_path}")
            return Response(json.dumps({"error": f"File {input_path} not found"}), mimetype='application/json'), 404
        if not input_path.endswith((".xls", ".xlsx")):
            logger.error(f"Unsupported file format: {input_path}")
            return Response(json.dumps({"error": "Only .xls or .xlsx files are supported"}), mimetype='application/json'), 400

        output_path = f"/data/{os.path.splitext(file_name)[0]}.csv"
        logger.info(f"Converting {input_path} to {output_path}")

        expected_headers = [
            "Час оформлення", "Назва ПМО", "Тип", "Номер МД", "Дата", "Відправник", "ЕДРПОУ", "Одержувач", "№",
            "Код товару", "Опис товару", "Кр.торг.", "Кр.відпр.", "Кр.пох.", "Умови пост.", "Місце пост", "К-ть",
            "Один.вим.", "Брутто, кг.", "Нетто, кг.", "Вага по МД", "ФВ вал.контр", "Особ.перем.", "43", "43_01",
            "РФВ Дол/кг.", "Вага.один.", "Вага різн.", "Контракт", "3001", "3002", "9610", "Торг.марк.",
            "РМВ Нетто Дол/кг.", "РМВ Дол/дод.од.", "РМВ Брутто Дол/кг", "Призн.Зед", "Мін.База Дол/кг.",
            "Різн.мін.база", "КЗ Нетто Дол/кг.", "КЗ Дол/шт.", "Різн.КЗ Дол/кг", "Різ.КЗ Дол/шт",
            "КЗ Брутто Дол/кг.", "Різ.КЗ Брутто", "пільгова", "повна"
        ]

        wb = load_workbook(input_path, read_only=True)
        sheet = wb[wb.sheetnames[0]]

        with open(output_path, 'w', encoding='utf-8', newline='') as csvfile:
            writer = csv.writer(csvfile)
            row_count = 0

            first_row = next(sheet.rows)
            headers = [cell.value for cell in first_row if cell.value is not None]
            headers = list(dict.fromkeys(headers))
            if not headers or len(headers) != len(expected_headers):
                logger.warning(f"Invalid or duplicate headers detected, using expected headers: {headers}")
                writer.writerow(expected_headers)
            else:
                writer.writerow(headers)

            seen_rows = set()
            first_data_row = True

            for row in sheet.rows:
                row_data = [cell.value for cell in row if cell.value is not None]
                if not row_data or all(v is None for v in row_data):
                    continue
                row_tuple = tuple(row_data)
                if row_tuple not in seen_rows:
                    if first_data_row and headers == expected_headers:
                        first_data_row = False
                    else:
                        writer.writerow(row_data)
                        seen_rows.add(row_tuple)
                        row_count += 1
                else:
                    logger.warning(f"Duplicate row detected: {row_data}")
                if row_count % 10000 == 0:
                    logger.info(f"Processed {row_count} unique rows")

        wb.close()
        logger.info(f"Total unique rows converted: {row_count}")

        return send_file(
            output_path,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f"{os.path.splitext(file_name)[0]}.csv"
        )
    except Exception as e:
        logger.error(f"Error in convert_excel_to_csv_from_data: {e}\n{traceback.format_exc()}")
        return Response(json.dumps({"error": str(e)}), mimetype='application/json'), 500

@app.route('/health', methods=['GET'])
def health_check():
    return Response(json.dumps({"status": "ok"}), mimetype='application/json'), 200

# Добавим диагностический эндпоинт для проверки совместимости
@app.route('/api/compatibility', methods=['GET'])
def check_compatibility():
    sample_response = {
        "id": "chat-12345",
        "object": "chat.completion",
        "created": int(time.time()),
        "model": "mistral:latest",
        "choices": [{
            "index": 0,
            "message": {
                "role": "assistant",
                "content": "Это тестовый ответ для проверки совместимости с Open WebUI"
            },
            "finish_reason": "stop"
        }]
    }
    return jsonify(sample_response)

if __name__ == "__main__":
    logger.info(f"Running Flask app on port {API_PORT}...")
    app.run(host="0.0.0.0", port=API_PORT, debug=True)